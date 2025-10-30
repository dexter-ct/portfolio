import sys
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.borders import Border, Side

# Get the CSV file path from command-line argument
if len(sys.argv) < 2:
    print("Usage: python SWIPE2.py <path_to_csv_file>")
    sys.exit(1)

csv_file = sys.argv[1]

# Check if the file exists
if not os.path.isfile(csv_file):
    print(f"Error: File not found - {csv_file}")
    sys.exit(1)

# Step 1: Read the CSV file while preserving leading zeros
df = pd.read_csv(csv_file, dtype=str)

# Step 2: Modify headers for columns F and G
if df.shape[1] >= 7:
    df.columns.values[5] = "     "  # Column F
    df.columns.values[6] = "      "  # Column G

# Step 3: Save to XLSX format in the same folder
base_name = os.path.splitext(os.path.basename(csv_file))[0]
xlsx_file = os.path.join(os.path.dirname(csv_file), base_name + ".xlsx")
df.to_excel(xlsx_file, index=False)

# Step 4: Apply styling using openpyxl
workbook = load_workbook(xlsx_file)
sheet = workbook.active

# Define styles
header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
header_font = Font(color="FFFFFF", bold=False)  # Ensure header is not bold
bold_font = Font(bold=True)
no_border = Border(left=Side(border_style=None),
                   right=Side(border_style=None),
                   top=Side(border_style=None),
                   bottom=Side(border_style=None))

# Style header row (A1 to H1)
for col in range(1, min(9, sheet.max_column + 1)):  # Columns A to H
    cell = sheet.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.border = no_border

# Style column H (excluding header)
if sheet.max_column >= 8:
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=8)  # Column H
        cell.font = bold_font

# Step 5: Save the styled workbook with today's date in dd.mm.yyyy format
today_str = datetime.today().strftime('%d.%m.%Y')
styled_xlsx_file = os.path.join(os.path.dirname(csv_file), today_str + ".xlsx")
workbook.save(styled_xlsx_file)

print(f"Styled Excel file saved as: {styled_xlsx_file}")
